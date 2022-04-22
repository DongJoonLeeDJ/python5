from openpyxl import Workbook,load_workbook

class MMy:
    def __init__(self):
        print("생성자")

    def save(self,cell,value):
        wb = Workbook()

        ws = wb.active
        ws[cell] = value

        wb.save('./ex0407/ex0407.xlsx')
        wb.close()
    
    def load(self,cell,value):
        wb = load_workbook('./ex0407/ex0407.xlsx')
        ws = wb.active

        # print(ws[cell].value)
        load_txt = ws[cell].value
        value.setText(load_txt)

        wb.close()
