from hashlib import new
from openpyxl import Workbook

def domake(filename):
    wb = Workbook()

    ws = wb.create_sheet()

    ws.title= "firstsheet"
    ws.sheet_properties.tabColor = "66e0ff"

    wb.create_sheet("ffsheet")
    wb.create_sheet("ccsheet",1)

    new_ws = wb["ccsheet"]

    print(wb.sheetnames)

    new_ws['A1'] ='qweqwe'
    new_ws['B1'] ='asdfasdf'

    target = wb.copy_worksheet(new_ws)
    target.title = "copied sheet"

    wb.save(filename)
    wb.close()