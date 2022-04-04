from openpyxl import Workbook,load_workbook

class Ex:
    def __init__(self):
        pass

    def doA1(self):
        wb = Workbook()

        ws = wb.active
        ws["C1"] = "aaa"
        ws["D1"] = "bbb"

        print(ws["C1"].value)
        print(ws["D1"].value)

        print(ws.cell(column=1,row=1).value)

        index = 1
        for x in range(1,11):
            for y in range(1,11):
                ws.cell(row=x,column=y,value=index)
                index += 1

        # self.doAA(b=20,a=10)

        wb.save("./ex0404/ex0404.xlsx")
        wb.close()


    def doB1(self):
        wb = load_workbook('./ex0404/ex0404.xlsx')
        ws = wb.active

        for x in range(1,11):
            for y in range(1,11):
                print(ws.cell(row=x,column=y).value,end="\t")
            print(end="\n")
        
        wb.close()
'''
    ctrl + / 주석 처리
'''
    # def doAA(self,a,b):
    #     print("a = ", a)
    #     print("b = ", b)